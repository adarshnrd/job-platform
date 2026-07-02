"""Export router — CSV, Excel, and Job Tracker downloads."""
from auth import get_user_id
import csv
import io
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from database import get_db
from services.job_tracker import get_tracker_stream

router = APIRouter(prefix="/export", tags=["export"])

db = get_db()


EXPORT_FIELDS = [
    ("job_title", "Job Title"),
    ("job_company", "Company"),
    ("job_location", "Location"),
    ("job_work_mode", "Work Mode"),
    ("source_platform", "Platform"),
    ("match_score", "Match Score"),
    ("match_tier", "Match Tier"),
    ("status", "Status"),
    ("applied_at", "Applied At"),
    ("interview_date", "Interview Date"),
    ("offered_salary", "Offered Salary"),
    ("notes", "Notes"),
    ("source_url", "Job URL"),
    ("created_at", "Discovered At"),
]


@router.get("/applications/csv")
async def export_applications_csv(
    user_id: str = Depends(get_user_id),
    status: str = Query(None),
):
    """Download all applications as a CSV file."""
    q = db.table("application_details").select("*").eq("user_id", user_id).order("created_at", desc=True)
    if status:
        q = q.eq("status", status)
    result = q.execute()
    apps = result.data or []

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([label for _, label in EXPORT_FIELDS])

    for app in apps:
        row = []
        for field, _ in EXPORT_FIELDS:
            val = app.get(field, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            row.append(str(val) if val is not None else "")
        writer.writerow(row)

    output.seek(0)
    filename = f"applications_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/applications/excel")
async def export_applications_excel(
    user_id: str = Depends(get_user_id),
    status: str = Query(None),
):
    """Download all applications as an Excel (.xlsx) file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    q = db.table("application_details").select("*").eq("user_id", user_id).order("created_at", desc=True)
    if status:
        q = q.eq("status", status)
    result = q.execute()
    apps = result.data or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Header row with styling
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="F59E0B", bold=True)

    headers = [label for _, label in EXPORT_FIELDS]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Status color map
    status_colors = {
        "applied": "22C55E", "interview_scheduled": "3B82F6",
        "offer_received": "F59E0B", "accepted": "10B981",
        "rejected": "EF4444", "withdrawn": "6B7280",
        "matched": "8B5CF6", "queued": "F97316",
    }

    for app in apps:
        row = []
        for field, _ in EXPORT_FIELDS:
            val = app.get(field, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            row.append(val if val is not None else "")
        ws.append(row)

        # Color-code by status
        status_val = app.get("status", "")
        color = status_colors.get(status_val, "FFFFFF")
        for col_idx in range(1, len(EXPORT_FIELDS) + 1):
            ws.cell(row=ws.max_row, column=col_idx).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )

    # Auto-size columns
    for col_idx, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    # Freeze header
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"applications_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([output.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/job-tracker")
async def download_job_tracker(user_id: str = Depends(get_user_id)):
    """Download the Job Tracker spreadsheet.

    Contains all discovered jobs with >= 50% match score. Rebuilt from
    the database on each request so it always reflects the latest state.
    """
    try:
        stream = get_tracker_stream(user_id)
        filename = f"job_tracker_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            iter([stream.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate tracker: {e}")
