"""
Job Tracker — maintains a persistent Excel spreadsheet of all discovered jobs.

Auto-updated after each discovery cycle and application attempt.
Only records jobs with match_score >= 50%.
Deduplicates by job_listing_id so re-runs update existing rows.
"""
import io
import os
from datetime import datetime, timezone
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db

MIN_TRACK_SCORE = 50

COLUMNS = [
    ("job_title",         "Job Title",           30),
    ("job_company",       "Company",             22),
    ("job_location",      "Location",            18),
    ("source_url",        "Job URL",             45),
    ("match_score",       "Match %",             10),
    ("status",            "Application Status",  18),
    ("failure_reason",    "Failure Reason",      25),
    ("created_at",        "Discovered At",       20),
    ("applied_at",        "Applied At",          20),
    ("source_platform",   "Source Platform",     15),
    ("job_work_mode",     "Work Mode",           12),
    ("match_tier",        "Match Tier",          13),
    ("job_required_skills", "Required Skills",   35),
    ("salary_range",      "Salary Range",        18),
    ("apply_url",         "Apply URL",           45),
    ("job_listing_id",    "Job ID",              12),
]

STATUS_DISPLAY = {
    "discovered": "Not Applied",
    "matched": "Not Applied",
    "queued": "Queued",
    "applying": "Applying",
    "applied": "Applied",
    "under_review": "Applied",
    "assessment": "Applied",
    "interview_scheduled": "Applied",
    "technical_round": "Applied",
    "hr_round": "Applied",
    "offer_received": "Applied",
    "accepted": "Applied",
    "rejected": "Rejected",
    "withdrawn": "Withdrawn",
}

STATUS_COLORS = {
    "Applied":      "22C55E",
    "Not Applied":  "94A3B8",
    "Queued":       "F97316",
    "Applying":     "3B82F6",
    "Failed":       "EF4444",
    "Rejected":     "EF4444",
    "Withdrawn":    "6B7280",
}

HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
HEADER_FONT = Font(color="F59E0B", bold=True, size=11)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="333333"),
    right=Side(style="thin", color="333333"),
)


def _tracker_path(user_id: str) -> str:
    data_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "trackers")
    os.makedirs(data_dir, exist_ok=True)
    return os.path.join(data_dir, f"job_tracker_{user_id[:8]}.xlsx")


def _format_dt(val) -> str:
    if not val:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    s = str(val)
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return s[:19]


def _row_from_app(app: dict) -> list:
    """Build a spreadsheet row from an application_details record."""
    raw_status = app.get("status", "matched")
    failure_reason = app.get("failure_reason") or ""

    if app.get("submission_status") == "failed":
        display_status = "Failed"
        failure_reason = failure_reason or app.get("submission_method", "")
    else:
        display_status = STATUS_DISPLAY.get(raw_status, raw_status.replace("_", " ").title())

    skills = app.get("job_required_skills") or []
    if isinstance(skills, list):
        skills = ", ".join(str(s) for s in skills)

    sal_min = app.get("salary_min")
    sal_max = app.get("salary_max")
    sal_cur = app.get("salary_currency") or ""
    if sal_min and sal_max:
        salary = f"{sal_cur} {sal_min:,}–{sal_max:,}"
    elif sal_min:
        salary = f"{sal_cur} {sal_min:,}+"
    else:
        salary = ""

    return [
        app.get("job_title", ""),
        app.get("job_company", ""),
        app.get("job_location", ""),
        app.get("source_url", ""),
        app.get("match_score", 0),
        display_status,
        failure_reason,
        _format_dt(app.get("created_at")),
        _format_dt(app.get("applied_at")),
        app.get("source_platform", ""),
        app.get("job_work_mode", ""),
        app.get("match_tier", ""),
        skills,
        salary,
        app.get("apply_url", ""),
        app.get("job_listing_id", ""),
    ]


def _create_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Tracker"

    headers = [label for _, label, _ in COLUMNS]
    ws.append(headers)

    for col_idx, (_, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    return wb


def _style_row(ws, row_num: int, status: str):
    color = STATUS_COLORS.get(status, "FFFFFF")
    status_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    status_col = next(i for i, (f, _, _) in enumerate(COLUMNS, 1) if f == "status")
    ws.cell(row=row_num, column=status_col).fill = status_fill
    ws.cell(row=row_num, column=status_col).font = Font(color="FFFFFF", bold=True)

    for col_idx in range(1, len(COLUMNS) + 1):
        ws.cell(row=row_num, column=col_idx).border = THIN_BORDER
        ws.cell(row=row_num, column=col_idx).alignment = Alignment(vertical="center")


def update_tracker(user_id: str) -> str:
    """Rebuild the tracker spreadsheet from the database. Returns the file path."""
    db = get_db()
    result = (
        db.table("application_details")
        .select("*")
        .eq("user_id", user_id)
        .gte("match_score", MIN_TRACK_SCORE)
        .order("match_score", desc=True)
        .execute()
    )
    apps = result.data or []

    wb = _create_workbook()
    ws = wb.active

    for app in apps:
        row_data = _row_from_app(app)
        ws.append(row_data)
        _style_row(ws, ws.max_row, row_data[5])

    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=11)
    ws.cell(row=summary_row + 1, column=1, value="Total Jobs Tracked")
    ws.cell(row=summary_row + 1, column=2, value=len(apps))
    applied = sum(1 for a in apps if STATUS_DISPLAY.get(a.get("status", ""), "") == "Applied")
    ws.cell(row=summary_row + 2, column=1, value="Applied")
    ws.cell(row=summary_row + 2, column=2, value=applied)
    ws.cell(row=summary_row + 3, column=1, value="Not Applied (manual review)")
    ws.cell(row=summary_row + 3, column=2, value=len(apps) - applied)
    ws.cell(row=summary_row + 4, column=1, value="Last Updated")
    ws.cell(row=summary_row + 4, column=2, value=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))

    path = _tracker_path(user_id)
    wb.save(path)
    logger.info(f"Job tracker updated: {len(apps)} jobs for user {user_id[:8]}… → {path}")
    return path


def get_tracker_bytes(user_id: str) -> bytes:
    """Return the tracker as bytes, rebuilding if needed."""
    path = _tracker_path(user_id)
    if not os.path.exists(path):
        update_tracker(user_id)

    with open(path, "rb") as f:
        return f.read()


def get_tracker_stream(user_id: str) -> io.BytesIO:
    """Return the tracker as a BytesIO stream, rebuilding from DB each time."""
    update_tracker(user_id)
    path = _tracker_path(user_id)
    buf = io.BytesIO()
    with open(path, "rb") as f:
        buf.write(f.read())
    buf.seek(0)
    return buf
